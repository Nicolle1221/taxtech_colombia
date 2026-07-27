import io
import json
import os
import sys
from pathlib import Path

from anthropic import Anthropic
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from pypdf import PdfReader
from supabase import create_client

PDF_PATH = Path("documentos/exogena_ejemplo.pdf")
MODEL = "claude-sonnet-5"

SYSTEM_PROMPT_EXTRACCION = (
    "Eres un experto en información exógena de la DIAN (Colombia). El texto de entrada "
    "puede venir de un PDF nativo, de un PDF escaneado/generado a partir de una hoja de "
    "Excel exportada desde el portal de la DIAN, o de texto copiado y pegado directamente "
    "de Excel. En esos casos las filas y columnas pueden llegar desalineadas, con "
    "tabulaciones o espacios irregulares, celdas partidas en varias líneas, o encabezados "
    "repetidos. Ignora ese ruido de formato y mapea los campos por su significado, no por "
    "su posición exacta en la línea:\n"
    "- 'concepto': el código numérico del concepto de la DIAN (ej. 5001).\n"
    "- 'valor_ingreso': el valor del pago o ingreso reportado por el tercero.\n"
    "- 'valor_retencion': el valor de la retención en la fuente asociada a ese concepto.\n"
    "Si un valor numérico no aparece explícitamente para un registro, usa 0 en vez de "
    "inventar una cifra.\n"
    "El campo 'datos_fiscales' es OBLIGATORIO en tu respuesta: inclúyelo siempre, incluso "
    "si el texto no trae ningún registro identificable de conceptos/valores, en cuyo caso "
    "debes enviar una lista vacía [] en vez de omitir el campo por completo.\n\n"
    "FORMATO 'Consulta de información reportada por terceros' (el reporte consolidado que "
    "cualquier contribuyente descarga de su cuenta en el portal de la DIAN, distinto de un "
    "certificado Formato 1001): cada fila trae 'NIT' y 'Nombre/Razón Social' del informante, "
    "un 'Detalle' con el código de concepto incrustado entre paréntesis (ej. 'Otros ingresos "
    "(Concepto: 5016)' -> concepto=5016), un único 'Valor', y una clasificación por 'Tope' "
    "(Tope 1: Ingresos, Tope 2: Patrimonio, Tope 3: Consumos TC, Tope 4: Movimientos/"
    "Consignaciones, Tope 5: Compras) o una etiqueta 'R30 Deudas'. Este formato NO trae "
    "retención en la fuente por separado: usa valor_retencion=0 para sus registros. "
    "Extrae a 'datos_fiscales' ÚNICAMENTE las filas marcadas como 'Tope 1: Ingresos' — son "
    "las únicas que representan renta gravable. Las filas de Tope 2 (Patrimonio), Tope 3 "
    "(Consumos TC), Tope 4 (Movimientos/Consignaciones), Tope 5 (Compras) y 'R30 Deudas' NO "
    "son ingresos: son saldos bancarios, consumos con tarjeta o deudas, y clasificarlas como "
    "ingreso sería un error grave. Ignóralas para 'datos_fiscales' aunque tengan un valor "
    "numérico grande.\n\n"
    "CERTIFICADO DE INGRESOS Y RETENCIONES (Formulario 220): si el documento es de este tipo, "
    "revisa quién lo emite:\n"
    "- Si el emisor es una entidad de las Fuerzas Militares o la Policía Nacional de Colombia "
    "(Ministerio de Defensa, Ejército, Armada, Fuerza Aérea, Policía Nacional, o similar), "
    "busca el 'exceso de salario básico' certificado (la diferencia entre el salario básico y "
    "la asignación salarial total, o cualquier rubro explícitamente etiquetado como 'exceso "
    "de salario básico' o equivalente) y repórtalo en el campo 'exceso_salario_basico_fuerza_"
    "publica'.\n"
    "- Si el emisor es una universidad PÚBLICA/oficial/estatal (no privada) y el certificado "
    "es de un rector o profesor, busca los 'gastos de representación' certificados y repórtalos "
    "en 'gastos_representacion_docente_publico'.\n"
    "- Si ninguno de estos dos casos aplica, deja ambos campos en 0. No inventes estos valores "
    "a partir de otro tipo de certificado o entidad."
)


class DatoFiscal(BaseModel):
    concepto: str
    descripcion_concepto: str
    nit_informante: str
    razon_social_informante: str
    valor_ingreso: float
    valor_retencion: float
    ano_gravable: int


class ContribuyenteExogena(BaseModel):
    nit: str
    nombre: str
    datos_fiscales: list[DatoFiscal] = Field(default_factory=list)
    exceso_salario_basico_fuerza_publica: float = 0.0
    gastos_representacion_docente_publico: float = 0.0


def _texto_de_reader(reader: PdfReader) -> str:
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def extraer_texto_pdf(ruta: Path) -> str:
    return _texto_de_reader(PdfReader(str(ruta)))


def extraer_texto_pdf_bytes(contenido: bytes) -> str:
    return _texto_de_reader(PdfReader(io.BytesIO(contenido)))


def leer_texto_plano(contenido: bytes) -> str:
    try:
        return contenido.decode("utf-8")
    except UnicodeDecodeError:
        return contenido.decode("latin-1")


def extraer_texto_xlsx_bytes(contenido: bytes) -> str:
    libro = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    lineas = []
    for hoja in libro.worksheets:
        lineas.append(f"--- Hoja: {hoja.title} ---")
        for fila in hoja.iter_rows(values_only=True):
            celdas = [str(valor) for valor in fila if valor is not None]
            if celdas:
                lineas.append("\t".join(celdas))
    return "\n".join(lineas)


def texto_a_estructura(client: Anthropic, texto: str) -> ContribuyenteExogena:
    response = client.messages.create(
        model=MODEL,
        max_tokens=2048,
        system=SYSTEM_PROMPT_EXTRACCION,
        tools=[
            {
                "name": "registrar_contribuyente_exogena",
                "description": "Registra los datos estructurados de información exógena de un contribuyente.",
                "input_schema": ContribuyenteExogena.model_json_schema(),
            }
        ],
        tool_choice={"type": "tool", "name": "registrar_contribuyente_exogena"},
        messages=[
            {
                "role": "user",
                "content": (
                    "Extrae la información exógena estructurada del siguiente texto "
                    f"proveniente de un reporte DIAN:\n\n{texto}"
                ),
            }
        ],
    )
    for block in response.content:
        if block.type == "tool_use":
            return ContribuyenteExogena.model_validate(block.input)
    raise RuntimeError("El modelo no devolvió una respuesta estructurada")


def guardar_en_supabase(datos: ContribuyenteExogena) -> None:
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL y SUPABASE_KEY deben estar definidas")
    client = create_client(url, key)
    client.table("contribuyentes_exogena").insert(
        json.loads(datos.model_dump_json())
    ).execute()


def main() -> None:
    if not PDF_PATH.exists():
        print(f"ERROR: no se encontró {PDF_PATH}", file=sys.stderr)
        sys.exit(1)

    texto = extraer_texto_pdf(PDF_PATH)
    client = Anthropic()
    datos = texto_a_estructura(client, texto)
    guardar_en_supabase(datos)
    print(f"OK: contribuyente {datos.nit} guardado en contribuyentes_exogena")


if __name__ == "__main__":
    main()
